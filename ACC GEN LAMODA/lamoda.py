# -*- coding: utf-8 -*-

from selenium import webdriver
import time
import requests
from selenium.common.exceptions import NoSuchElementException
from openpyxl import load_workbook
import random
import string


def generate_random_string(length):
    letters = string.ascii_lowercase
    rand_string = ''.join(random.choice(letters) for _ in range(length))
    return rand_string


out = open('lamoda.txt', 'w')
wb = load_workbook('data.xlsx')
sheet = wb['info']
retry = int(sheet.cell(row=8, column=2).value)
token = str(sheet.cell(row=1, column=2).value)
domain = str(sheet.cell(row=2, column=2).value)
password = str(sheet.cell(row=3, column=2).value)
name = str(sheet.cell(row=4, column=2).value)
surname = str(sheet.cell(row=5, column=2).value)
middle_name = str(sheet.cell(row=6, column=2).value)
date = str(sheet.cell(row=7, column=2).value)
date = date[8] + date[9] + '.' + date[5] + date[6] + '.' + date[:4]
c = 0

while c < retry:
    c += 1
    mail = generate_random_string(6) + '@' + domain
    driver = webdriver.Chrome()
    driver.maximize_window()
    reg_type = 0

    driver.get('https://www.lamoda.ru/')
    time.sleep(2)


    def check_exists_by_xpath(xpath):
        try:
            driver.find_element_by_xpath(xpath)
        except NoSuchElementException:
            return False
        return True


    # Войти
    driver.find_element_by_xpath("/html/body/div[1]/div[2]/div/div/a[1]").click()
    time.sleep(2)

    # Рега
    if check_exists_by_xpath("/html/body/div[1]/div[6]/div[6]/div/div/div[2]/div[2]/div/div[2]/div/div[2]/div/div[1]/form/div[3]/a") or\
            check_exists_by_xpath('/html/body/div[1]/div[6]/div[7]/div[1]/div/div[2]/div[2]/div/div[2]/div/div[2]/div/div[1]/form/div[3]/a'):
        a = driver.find_element_by_xpath('/html/body/div[1]/div[6]/div[6]/div/div/div[2]/div[2]/div/div[2]/div/div[2]/div/div[1]/form/div[3]/a')
        driver.execute_script("arguments[0].click();", a)
        reg_type = 1
    else:
        a = driver.find_element_by_xpath('/html/body/div[1]/div[6]/div[6]/div/div/div[2]/div[2]/div/div[2]/div/div[2]/div[1]/form/div[1]/span[2]/a')
        driver.execute_script("arguments[0].click();", a)
        reg_type = 2
    time.sleep(2)

    if reg_type == 1:
        # Почта
        driver.find_element_by_xpath("/html/body/div[1]/div[6]/div[7]/div[2]/div/div[3]/div[2]/div/div/div[1]/form/div[1]/div/div[1]/input").send_keys(mail)
        time.sleep(1)

        # Рассылка
        driver.find_element_by_xpath("/html/body/div[1]/div[6]/div[7]/div[2]/div/div[3]/div[2]/div/div/div[1]/form/div[2]/div[2]").click()
        time.sleep(1)

        # Телефон
        url = 'https://onlinesim.ru/api/getNum.php?apikey=' + token + '&service=lamoda'
        response = requests.get(url).json()
        tzid = str(response['tzid'])

        url = 'https://onlinesim.ru/api/getState.php?apikey=' + token + '&tzid=' + tzid
        response = requests.get(url).json()
        number = str(response[0]['number'][1:])

        driver.find_element_by_xpath("/html/body/div[1]/div[6]/div[7]/div[2]/div/div[3]/div[2]/div/div/div[1]/form/div[3]/div/input").send_keys(number)
        time.sleep(1)

        # Пароль
        driver.find_element_by_xpath("/html/body/div[1]/div[6]/div[7]/div[2]/div/div[3]/div[2]/div/div/div[1]/form/div[4]/div/input").send_keys(password)
        time.sleep(1)

        # Повтор пароля
        driver.find_element_by_xpath("/html/body/div[1]/div[6]/div[7]/div[2]/div/div[3]/div[2]/div/div/div[1]/form/div[5]/div/input").send_keys(password)
        time.sleep(1)

        # Имя
        driver.find_element_by_xpath("/html/body/div[1]/div[6]/div[7]/div[2]/div/div[3]/div[2]/div/div/div[1]/form/div[6]/div/input").send_keys(name)
        time.sleep(1)

        # Фамилия
        driver.find_element_by_xpath("/html/body/div[1]/div[6]/div[7]/div[2]/div/div[3]/div[2]/div/div/div[1]/form/div[7]/div/input").send_keys(surname)
        time.sleep(1)

        # Зарегистироваться
        driver.find_element_by_xpath("/html/body/div[1]/div[6]/div[7]/div[2]/div/div[3]/div[2]/div/div/div[1]/form/button").click()
        time.sleep(15)

        # Код
        check = 0
        msg = ""
        while check < 190:
            url = 'https://onlinesim.ru/api/getState.php?apikey=' + token + '&tzid=' + tzid
            response = requests.get(url).json()
            try:
                msg = str(response[0]['msg'])
            except:
                msg = ""
            if msg == "": check += 1
            else: break
        while msg == '':
            url = 'https://onlinesim.ru/api/getNum.php?apikey=' + token + '&service=lamoda'
            response = requests.get(url).json()
            tzid = str(response['tzid'])
        
            url = 'https://onlinesim.ru/api/getState.php?apikey=' + token + '&tzid=' + tzid
            response = requests.get(url).json()
            number = str(response[0]['number'][2:])
        
            check = 0
            msg = ""
            while check < 190:
                url = 'https://onlinesim.ru/api/getState.php?apikey=' + token + '&tzid=' + tzid
                response = requests.get(url).json()
                try:
                    msg = str(response[0]['msg'])
                except:
                    msg = ""
                if msg == "":
                    check += 1
                else:
                    break
        time.sleep(5)
        driver.find_element_by_xpath('/html/body/div[1]/div[6]/div[7]/div[3]/div/div[3]/div[2]/form/div[2]/div/div/div[1]/input').send_keys(msg)
        out.write('+' + number + ' ' + mail + ' ' + password + '\n')
        time.sleep(5)

        driver.get('https://www.lamoda.ru/customer/account/')

        # Отчество
        driver.find_element_by_xpath('/html/body/div[1]/div[5]/div/div/div[2]/div/div[2]/div[1]/div[3]/div/div[1]/input').send_keys(middle_name)
        time.sleep(1)

        # Дата рождения
        driver.find_element_by_xpath('/html/body/div[1]/div[5]/div/div/div[2]/div/div[2]/div[5]/div[2]/div/div[1]/input').send_keys(date)
        time.sleep(1)

        # Пол
        driver.find_element_by_xpath('//*[@id="vue-root"]/div/div/div[2]/div/div[2]/div[6]/div[2]/div/button[1]').click()
        time.sleep(1)

        # Сохранить
        driver.find_element_by_xpath('//*[@id="vue-root"]/div/div/div[2]/div/div[2]/div[7]/div[2]/button').click()
        time.sleep(5)

        driver.quit()

    if reg_type == 2:
        c -= 1
        driver.quit()
